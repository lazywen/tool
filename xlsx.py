#!/usr/bin/env python
# -*- coding: utf-8 -*-

import openpyxl

def getdata_from_xlsx(xlsx):
    ''' return {
            sheet1_name: (
                (a1, b1, c1, ...), #row1
                (a2, b2, c2, ...), #row2
                    ...
            ),
            ...
        }
    '''
    try:
        wb = openpyxl.load_workbook(xlsx)
    except Exception,e:
        glog('error', 'can not read xlsx file: %s' % xlsx)
        raise e
    datas = {}
    for sheet in wb.worksheets:
        sheet_data = []
        for row in sheet.rows:
            row_value = [ cell.value for cell in row ]
            sheet_data.append(row_value)
        datas[sheet.title] = sheet_data
    return datas

def create_xlsx(path, datas):
    ''' datas = {
            sheet1_name: (
                (a1, b1, c1, ...), #row1
                (a2, b2, c2, ...), #row2
                    ...
            ),
            ...
        }
    '''
    from openpyxl.cell import get_column_letter
    wb = openpyxl.Workbook()
    first_sheet = True
    for sheet in datas:
        if first_sheet:
            ws = wb.get_active_sheet()
            first_sheet = False
        else:
            ws = wb.create_sheet()
        ws.title = sheet

        for row_index, row_data in enumerate(datas[sheet], start=1):
            for col_index, row_value in enumerate(row_data, start=1):
                col = get_column_letter(col_index)
                if type(row_value) == type(u''):
                    row_value = row_value.encode('utf-8')
                ws.cell('%s%s' % (col, row_index)).value = row_value
    try:
        wb.save(filename = path)
    except Exception,e:
        glog('error', 'can not save file to %s' % path)
        raise e
